#!/usr/bin/env python3
"""Import the user-supplied schedule JSON workbook into a compact reference file."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = ROOT / "references" / "manual" / "schedule_json_enrichment.json"
EVENT_DATES = ["2026-04-24", "2026-04-25", "2026-04-26"]


def text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def clean_text(value) -> str:
    value = text(value)
    lower = value.lower()
    placeholder_markers = [
        "⚠️",
        "这里添加",
        "add location here",
        "add host",
        "add description",
    ]
    if any(marker in lower for marker in placeholder_markers):
        return ""
    return value


def split_ids(value) -> list[str]:
    return [part for part in re.split(r"[,，\s]+", text(value)) if part]


def parse_dates(date_label: str, time_text: str) -> tuple[list[str], str]:
    combined = f"{date_label} {time_text}"
    days = sorted({int(day) for day in re.findall(r"(?<!\d)(2[4-6])\s*[日号]", combined)})
    if days:
        return [f"2026-04-{day:02d}" for day in days], "explicit"
    if "多日" in combined:
        return EVENT_DATES[:], "multi-day"
    return [], "unknown"


def normalize_host(host: dict) -> dict:
    return {
        "name": clean_text(host.get("name")),
        "bio": clean_text(host.get("bio")),
    }


def normalize_segment(segment: dict) -> dict:
    speakers = []
    for speaker in segment.get("speakers") or []:
        if not isinstance(speaker, dict):
            continue
        normalized = {
            "name": clean_text(speaker.get("name")),
            "bio": clean_text(speaker.get("bio")),
        }
        if normalized["name"] or normalized["bio"]:
            speakers.append(normalized)
    return {
        "title": clean_text(segment.get("title")),
        "speakers": speakers,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    workbook = load_workbook(args.source, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
    records = []

    for row_index in range(2, worksheet.max_row + 1):
        row = {
            str(headers[column - 1]): worksheet.cell(row_index, column).value
            for column in range(1, worksheet.max_column + 1)
        }
        raw_json = text(row.get("采纳内容"))
        if not raw_json:
            continue
        payload = json.loads(raw_json)
        ch = payload.get("ch") or {}
        en = payload.get("en") or {}
        date_label = text(row.get("日期"))
        time_text = text(row.get("时间"))
        date_tags, date_precision = parse_dates(date_label, time_text)
        hosts = [
            normalized
            for host in ch.get("hosts") or []
            if isinstance(host, dict)
            for normalized in [normalize_host(host)]
            if normalized["name"] or normalized["bio"]
        ]
        segments = [
            normalized
            for segment in ch.get("segments") or []
            if isinstance(segment, dict)
            for normalized in [normalize_segment(segment)]
            if normalized["title"] or normalized["speakers"]
        ]

        record = {
            "source_row": row_index,
            "original_title": text(row.get("原始标题")),
            "hidden": text(row.get("隐藏")),
            "note": text(row.get("备注")),
            "schedule_status": text(row.get("场次排期")),
            "activity_detail_ids": split_ids(row.get("活动详情")),
            "duration_units": text(row.get("时长单位")),
            "period_label": text(row.get("时间段")),
            "date_label": date_label,
            "date_tags": date_tags,
            "date_precision": date_precision,
            "time_text": time_text,
            "sort_group": text(row.get("顺序2")),
            "sort_number": text(row.get("顺序编号")),
            "hive_intent_id": text(row.get("🔒蜂巢意向表")),
            "board_title": text(row.get("新生论坛 热带雨林")),
            "location_raw": text(row.get("场地2")),
            "container": text(row.get("容器")),
            "venue_type": text(row.get("场地类型")),
            "done": text(row.get("搞定")),
            "legacy_detail_ids": split_ids(row.get("活动详情 副本041900")),
            "title": clean_text(ch.get("title")),
            "title_en": clean_text(en.get("title")),
            "location_from_json": clean_text(ch.get("location")),
            "hosts": hosts,
            "segments": segments,
            "description": clean_text(ch.get("description")),
            "source_level": "schedule-json",
        }
        records.append(record)

    output = {
        "schema_version": 1,
        "source": args.source.name,
        "source_note": "从用户补充的日程JSON工作簿抽取；作为官网活动表和公众号文章之外的结构化补充层使用。",
        "columns": [str(header) for header in headers],
        "records": records,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {args.output} records={len(records)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
